from PyQt5.QtWidgets import QMainWindow, QVBoxLayout, QWidget, QPushButton, QLabel, QComboBox, QMessageBox, QListWidget, QListWidgetItem, QHBoxLayout
from PyQt5.QtCore import Qt, QDate
import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter
import os

class GenRec(QMainWindow):
    def __init__(self, db):
        super(GenRec, self).__init__()
        self.db = db
        self.initUI()
        self.load_intervals()

    def initUI(self):
        self.setWindowTitle('Generar Archivo Excel')
        self.setGeometry(100, 100, 600, 400)  # Aumenta el tamaño para mostrar la lista

        layout = QVBoxLayout()

        self.list_widget = QListWidget(self)
        layout.addWidget(self.list_widget)

        self.central_widget = QWidget()
        self.central_widget.setLayout(layout)
        self.setCentralWidget(self.central_widget)

    def load_intervals(self):
        try:
            query = "SELECT folio, fecha_inicio, hora_inicio, fecha_fin, hora_fin FROM suma_historial_recaudo"
            self.db.cursor.execute(query)
            intervals = self.db.cursor.fetchall()

            self.list_widget.clear()
            for interval in intervals:
                folio, fecha_inicio, hora_inicio, fecha_fin, hora_fin = interval
                item_text = f"Folio: {folio}, Desde: {fecha_inicio} {hora_inicio}, Hasta: {fecha_fin} {hora_fin}"
                
                item_widget = QWidget()
                item_layout = QHBoxLayout()
                item_layout.setContentsMargins(0, 0, 0, 0)
                
                item_label = QLabel(item_text)
                item_label.setFixedHeight(25)
                
                generate_btn = QPushButton("Generar")
                generate_btn.setStyleSheet("background-color: rgb(127, 98, 184);")
                generate_btn.setFixedSize(60, 16)
                generate_btn.clicked.connect(lambda ch, interval=interval: self.generate_excel(interval))
                
                item_layout.addWidget(item_label)
                item_layout.addWidget(generate_btn)
                
                item_widget.setLayout(item_layout)
                
                list_item = QListWidgetItem(self.list_widget)
                list_item.setSizeHint(item_widget.sizeHint())
                self.list_widget.addItem(list_item)
                self.list_widget.setItemWidget(list_item, item_widget)
                
        except psycopg2.Error as e:
            QMessageBox.critical(self, 'Error', f'Error al cargar intervalos: {e}', QMessageBox.Ok)

    def generate_excel(self, interval=None):
        if interval is None:
            interval = self.list_widget.currentItem().data(Qt.UserRole)

        if not interval:
            QMessageBox.critical(self, 'Error', 'Seleccione un intervalo.', QMessageBox.Ok)
            return

        folio, fecha_inicio, hora_inicio, fecha_fin, hora_fin = interval

        try:
            query = """
            SELECT hr.folio, hr.fecha, hr.hora, hr.eco, 
                   hr.id_chofer1,
                   ch1.nombre || ' ' || ch1.apellido_paterno || ' ' || ch1.apellido_materno AS chofer_1,
                   hr.id_chofer2,
                   ch2.nombre || ' ' || ch2.apellido_paterno || ' ' || ch2.apellido_materno AS chofer_2,
                   hr.monedas, hr.billetes, hr.monedas + hr.billetes AS total
            FROM historial_recaudo hr
            LEFT JOIN empleado_chofer ch1 ON hr.id_chofer1 = ch1.id_chofer
            LEFT JOIN empleado_chofer ch2 ON hr.id_chofer2 = ch2.id_chofer
            WHERE (hr.fecha || ' ' || hr.hora)::timestamp BETWEEN %s AND %s
            """
            self.db.cursor.execute(query, (f"{fecha_inicio} {hora_inicio}", f"{fecha_fin} {hora_fin}"))
            records = self.db.cursor.fetchall()

            # Crear el archivo Excel
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Recaudo Intervalo'

            headers = ['Folio', 'Fecha', 'Hora', 'Eco', 'ID Chofer 1', 'Nombre Chofer 1', 'ID Chofer 2', 'Nombre Chofer 2', 'Monedas', 'Billetes', 'Total']
            sheet.append(headers)

            for record in records:
                sheet.append(record)

            # Calcular los totales
            total_monedas = sum(record[8] for record in records)
            total_billetes = sum(record[9] for record in records)
            total_recaudo = total_monedas + total_billetes

            # Añadir información del intervalo y totales al final del archivo
            sheet.append([])
            sheet.append(['Intervalo', f'Desde {fecha_inicio} {hora_inicio}', f'Hasta {fecha_fin} {hora_fin}'])
            sheet.append(['Total Monedas', total_monedas])
            sheet.append(['Total Billetes', total_billetes])
            sheet.append(['Total Recaudo', total_recaudo])

            for col in range(1, len(headers) + 1):
                sheet.column_dimensions[get_column_letter(col)].width = 15

            # Definir la ruta específica donde quieres guardar el archivo
            output_directory = r'C:\Users\Cesar\Desktop\Excel'  # Cambia esta línea a tu ruta deseada
            os.makedirs(output_directory, exist_ok=True)  # Crea el directorio si no existe

            # Guardar el archivo
            file_path = os.path.join(output_directory, f'Recaudo_Intervalo_Folio_{folio}.xlsx')
            workbook.save(file_path)
            QMessageBox.information(self, 'Éxito', f'Archivo Excel generado correctamente en {file_path}.', QMessageBox.Ok)

        except psycopg2.Error as e:
            QMessageBox.critical(self, 'Error', f'Error al generar el archivo Excel: {e}', QMessageBox.Ok)

        except Exception as e:
            QMessageBox.critical(self, 'Error', f'Error inesperado: {e}', QMessageBox.Ok)
